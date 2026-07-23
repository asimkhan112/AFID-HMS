"""
excel_exporter.py
Utility for generating Excel (.xlsx) exports of patient queue data.
"""

import os
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_filename(name: str) -> str:
    """Remove or replace characters that are invalid in filenames."""
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()


def generate_queue_excel(patients: List[Dict[str, Any]], doctor_name: str) -> str:
    """
    Generate an Excel file containing patient queue data.
    
    Args:
        patients: List of patient dictionaries with queue information
        doctor_name: Name of the doctor for the filename
        
    Returns:
        str: Path to the generated Excel file
    """
    # Create exports directory if it doesn't exist
    exports_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(exports_dir, exist_ok=True)
    
    # Generate filename: DoctorName_YYYY-MM-DD_HH-MM-SSS.xlsx (include seconds to avoid collisions)
    now = datetime.now()
    sanitized_doctor_name = sanitize_filename(doctor_name)
    filename = f"{sanitized_doctor_name}_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    
    # Note: We intentionally allow overwriting files from the same minute.
    # The logout endpoint may be called multiple times within the same minute,
    # and each call should update the same file with the current queue state.
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Queue Export"
    
    # Define headers
    headers = [
        "Queue Number",
        "Patient ID",
        "Patient Name",
        "Age",
        "Gender",
        "Visit Date",
        "Visit Time",
        "Status",
        "Doctor Name"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="053c2a", end_color="053c2a", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_border = Border(
        left=Side(style="thin", color="cbd5e1"),
        right=Side(style="thin", color="cbd5e1"),
        top=Side(style="thin", color="cbd5e1"),
        bottom=Side(style="thin", color="cbd5e1")
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write patient data
    for row_num, patient in enumerate(patients, 2):
        # Queue Number (row index)
        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = cell_alignment
        ws.cell(row=row_num, column=1).border = cell_border
        
        # Patient ID (mr_number)
        ws.cell(row=row_num, column=2, value=patient.get("mr_number", "")).alignment = cell_alignment
        ws.cell(row=row_num, column=2).border = cell_border
        
        # Patient Name
        ws.cell(row=row_num, column=3, value=patient.get("full_name", "")).alignment = cell_alignment
        ws.cell(row=row_num, column=3).border = cell_border
        
        # Age (not directly in model, using placeholder or calculated from registered_at if needed)
        # For now, leaving as N/A or could be calculated
        ws.cell(row=row_num, column=4, value=patient.get("age", "N/A")).alignment = cell_alignment
        ws.cell(row=row_num, column=4).border = cell_border
        
        # Gender
        ws.cell(row=row_num, column=5, value=patient.get("gender", "")).alignment = cell_alignment
        ws.cell(row=row_num, column=5).border = cell_border
        
        # Visit Date (from registered_at or session_date)
        visit_date = patient.get("visit_date") or patient.get("registered_at", "")
        if isinstance(visit_date, datetime):
            visit_date = visit_date.strftime("%Y-%m-%d")
        ws.cell(row=row_num, column=6, value=visit_date).alignment = cell_alignment
        ws.cell(row=row_num, column=6).border = cell_border
        
        # Visit Time
        visit_time = patient.get("visit_time") or patient.get("registered_at", "")
        if isinstance(visit_time, datetime):
            visit_time = visit_time.strftime("%H:%M")
        ws.cell(row=row_num, column=7, value=visit_time).alignment = cell_alignment
        ws.cell(row=row_num, column=7).border = cell_border
        
        # Status
        ws.cell(row=row_num, column=8, value=patient.get("status", "")).alignment = cell_alignment
        ws.cell(row=row_num, column=8).border = cell_border
        
        # Doctor Name
        ws.cell(row=row_num, column=9, value=doctor_name).alignment = cell_alignment
        ws.cell(row=row_num, column=9).border = cell_border
    
    # Auto-adjust column widths
    column_widths = [12, 15, 30, 8, 10, 12, 12, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(filepath)
    
    return filepath


def get_exported_file_path(doctor_name: str) -> str:
    """
    Get the expected filepath for a doctor's queue export.
    Used to check if a file was created for the current minute.
    
    Args:
        doctor_name: Name of the doctor
        
    Returns:
        str: Expected filepath for the export
    """
    exports_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(exports_dir, exist_ok=True)
    
    now = datetime.now()
    sanitized_doctor_name = sanitize_filename(doctor_name)
    filename = f"{sanitized_doctor_name}_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return os.path.join(exports_dir, filename)
